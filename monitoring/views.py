import re
from datetime import datetime, date, timedelta
import json
from pathlib import Path
from zoneinfo import ZoneInfo


from django.http import HttpResponseRedirect, HttpResponseForbidden, JsonResponse, HttpResponse
from django.conf import settings
from django.shortcuts import render
from django.utils import timezone
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

import pyodbc
from urllib.parse import urlencode
from django.urls import reverse

from .utils import can_edit_from_request
from monitoring_PTC.charts.http_clients import fetch_xml
from monitoring_PTC.charts.xml_parser import parse_series
from monitoring_PTC.charts.timezone_utils import TZ_CHISINAU, to_epoch_seconds
from concurrent.futures import ThreadPoolExecutor, as_completed

# --- Конфиг таймаута подключения к SQL Server (секунды) ---
DB_CONNECT_TIMEOUT = getattr(settings, "DB_CONNECT_TIMEOUT", 5)

# --- Simple memory cache for LR requests ---
LR_MEMORY_CACHE = {}  # {(ips, id_param): (value, ts)}
LR_CACHE_TTL = 300  # seconds

# Включать ли фолбэк на param_rokura, если в IDS нет id_lovati для нужного параметра
# ВАЖНО: для LOVATI фолбэк отключаем, чтобы при отсутствии корректного ID ячейка была пустой.
LOVATI_FALLBACK_TO_PARAM = False

# LOVATI: объекты, где PTI.pompa/pompa2/pompa3 = 0/1 (0=ON зелёный, 1=OFF красный).
# Здесь насосы НЕ должны зависеть от IDS/LR.
LOVATI_PUMP01_PTC = {
    "2113","2209","2407","1012","1018","3001","3002","3003","3004","3005",
    "3013","3025","3038","3043","3054","3064","3083","3107","3111","3118",
    "3127","4018","4044","5012","5013","5023","5043","5051",
}



# -------- Local JSON storage for exclusions & comments --------
STORAGE_DIR = Path(settings.BASE_DIR) / "storage"
STORAGE_DIR.mkdir(parents=True, exist_ok=True)
EXCLUSIONS_PATH = STORAGE_DIR / "exclusions.json"
COMMENTS_PATH = STORAGE_DIR / "comments.json"

# --- Исключения: маппинг «человеческое имя» <-> внутренний ключ ---
HUMAN_TO_KEY = {
    "Q": "q1",
    "G1": "g1",
    "G2": "g2",
    "ΔG": "dg",
    "Δ%": "dg_pct",
    "T1": "t1",
    "T2": "t2",
    "ΔT": "dt",
    "T31": "t31",
    "T32": "t32",
    "T41": "t41",
    "T42": "t42",
    "T43": "t43",
    "T44": "t44",
    "Tacm": "tacm",
    "Gacm": "gacm",
    "Gacm-P": "gacm_p",
    "ΔGacm": "dgacm_val",
    "Gadaos": "g_adaos",
    "220V": "sursa",
    "Pompa": "pompa",
    "Data/Ora": "time",
    # Спец-метка для исключения всего объекта
    "PTC (tot)": "__all__",
}
KEY_TO_HUMAN = {v: k for k, v in HUMAN_TO_KEY.items()}

# Цветовые схемы для Excel под светлую / тёмную тему
EXCEL_THEME_COLORS = {
    "light": {
        # шапка
        "header_bg": "FF7291BF",  # var(--accent) #7291bf
        "header_font": "FFFFFFFF",  # белый текст

        # строки
        "row_even_bg": "FFF8F9FA",
        "row_odd_bg": "FFFFFFFF",

        # текст
        "text": "FF1F1F29",  # var(--text) #1f1f29

        # аварии / стутусы
        "danger": "FFD21919",  # красный (T2, T4, Tacm и т.п.)
        "good": "FF0BC63D",  # зелёный (220V OK, помпа OK)
        "voltage_off": "FFE03131",  # красный для 220V OFF

        # доп. цвета для помп
        "pump_yellow": "FFE7D000",  # остановлена / LCS < 30
        "pump_gray": "FFB0B0B0",  # нет тока / неизвестно
    },
    "dark": {
        # шапка (как верхняя часть градиента)
        "header_bg": "FF2C3034",
        "header_font": "FFFFFFFF",

        # строки
        "row_even_bg": "FF3C4349",
        "row_odd_bg": "FF343A40",

        # текст
        "text": "FFF8F9FA",

        # аварии / стутусы
        "danger": "FFFF6666",  # чуть ярче в тёмной теме
        "good": "FF00FB29",
        "voltage_off": "FFE03131",

        # помпы — те же цвета, чтобы было узнаваемо
        "pump_yellow": "FFE7D000",
        "pump_gray": "FFB0B0B0",
    },
}


def _norm_excl_param(param: str) -> list[str]:
    """
    Приводит значение 'param' из exclusions.json к внутренним ключам.
    Возвращает список ключей (на случай старого 't4' → ['t41','t42','t43','t44']).
    Поддерживает как "человеческие" ярлыки (например, 'Δ%'), так и внутренние ключи.
    Дополнительно поддерживает исключение всего объекта: '__all__' / 'PTC (tot)' / 'tot' / 'all' / '*'.
    """
    p = (param or "").strip()
    if not p:
        return []

    # поддержка «человеческих» ярлыков
    if p in HUMAN_TO_KEY:
        p = HUMAN_TO_KEY[p]

    pl = p.lower()
    # исключение всей строки PTC
    if pl in ("__all__", "ptc (tot)", "ptc", "tot", "all", "*"):
        return ["__all__"]

    # обратная совместимость со старым 't4' → раскладываем в конкретные T4*
    if pl == "t4":
        return ["t41", "t42", "t43", "t44"]

    return [p]


# --- безопасная работа с JSON ---
def _load_json(path: Path) -> dict:
    """
    Безопасно читает JSON; при ошибке или пустом файле возвращает {}.
    """
    if not path.exists():
        return {}
    try:
        txt = path.read_text(encoding="utf-8")
        if not txt.strip():
            return {}
        return json.loads(txt)
    except (OSError, json.JSONDecodeError):
        return {}


def _save_json(path: Path, data: dict) -> None:
    """
    Безопасная запись JSON: создаём каталог и пишем во временный файл с последующей заменой.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    tmp_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(path)


# --- Активные исключения из JSON (storage/exclusions.json) ---
def _get_active_exclusions(today: date, tura_cur: str = "1") -> dict[str, set[str]]:
    """
    Вернёт { '5118': {'t1','t2','dg','dg_pct','t41','time','__all__',...}, ... }
    Исключение активно, если текущее локальное время попадает в интервал [start, end] либо,
    для старых записей, если дата «until» ≥ today.
    """
    raw = _load_json(EXCLUSIONS_PATH) or {}
    active: dict[str, set[str]] = defaultdict(set)

    # Текущее время в зоне Кишинёва (учитывает лето/зиму)
    now_dt = timezone.now().astimezone(TZ_CHISINAU)

    for ptc, items in raw.items():
        it_list = items or []
        for it in it_list:
            # читаем строки start/end из JSON
            start_str = str(it.get("start") or "").strip()
            end_str   = str(it.get("end") or "").strip()

            is_active = False
            if start_str or end_str:
                # разбираем строки в datetime; None, если пусто или ошибка
                try:
                    start_dt = datetime.fromisoformat(start_str) if start_str else None
                except Exception:
                    start_dt = None
                try:
                    end_dt   = datetime.fromisoformat(end_str) if end_str else None
                except Exception:
                    end_dt   = None

                # определяем начало и конец интервала в локальной зоне:
                check_start = True
                if start_dt:
                    # если дата без зоны, считаем её временем Кишинёва
                    if start_dt.tzinfo is None:
                        start_dt = start_dt.replace(tzinfo=TZ_CHISINAU)
                    # интервал активен только после старта
                    check_start = now_dt >= start_dt

                check_end = True
                if end_dt:
                    if end_dt.tzinfo is None:
                        end_dt = end_dt.replace(tzinfo=TZ_CHISINAU)
                    # интервал активен только до конца
                    check_end = now_dt <= end_dt

                is_active = check_start and check_end
            else:
                # для старых записей: только поле until (дата без времени)
                until = str(it.get("until") or "").strip()
                try:
                    is_active = date.fromisoformat(until) >= today
                except Exception:
                    is_active = False

            # если интервал не активен, пропускаем этот параметр
            if not is_active:
                continue

            # иначе добавляем все внутренние ключи этого параметра
            params = _norm_excl_param(str(it.get("param") or ""))
            for p in params:
                if p:
                    active[str(ptc)].add(p)

    return active



# ───────── helpers ─────────
def _dsn(dct) -> str:
    """
    Строит DSN-строку для pyodbc из словаря настроек.
    Пропускаем пустые значения и гарантируем TrustServerCertificate=yes.
    """
    parts = [f"{k}={v}" for k, v in dct.items() if v]
    if not any(p.startswith("TrustServerCertificate=") for p in parts):
        parts.append("TrustServerCertificate=yes")
    return ";".join(parts)


def _to_float(x, default=0.0):
    try:
        if x is None:
            return default
        if isinstance(x, str):
            s = x.replace(",", ".").strip()
            if s == "":
                return default
            return float(s)
        return float(x)
    except Exception:
        return default


def _roundf(x, nd=2):
    try:
        return round(_to_float(x), nd)
    except Exception:
        return 0.0

def _disp_num(x, nd=2):
    """
    Для таблицы мониторинга:
    - None  -> ''
    - 0     -> 0
    - число -> округлённое число
    """
    if x is None:
        return ""
    try:
        return round(float(x), nd)
    except Exception:
        return ""



def _load_gacm_hourly_template(
    ptc_codes: list[str],
    day: date,
    typeobj_filter: str | None = "= 0",
) -> dict[str, dict[int, float]]:
    """Загружает почасовые прогнозы GacmPredictPTC (LOVATI) в виде:
    { '1013': {0: 0.12, 1: 0.11, ...}, ... }

    Связь: LOVATI.dbo.PTI.id -> LOVATI.dbo.GacmPredictPTC.pti

    typeobj_filter:
      - "= 0"  : только LOVATI PTC (как раньше)
      - None    : без фильтра typeObj (нужно для TERMOCOM PTC, которые тоже имеют typeObj=0)
      - "<> 0" : если вдруг нужно отбирать non-zero typeObj
    """
    if not ptc_codes:
        return {}

    ptc_codes = sorted({(c or "").strip() for c in ptc_codes if c})
    if not ptc_codes:
        return {}

    dsn_lovati = _dsn(settings.LOVATI_SERVER)
    placeholders = ",".join("?" for _ in ptc_codes)

    day_start = datetime(day.year, day.month, day.day, 0, 0, 0)
    day_end = day_start + timedelta(days=1)

    type_cond = f"AND p.typeObj {typeobj_filter}" if typeobj_filter else ""

    sql = rf"""
        WITH src AS (
            SELECT
                RTRIM(p.pti)                           AS ptc_code,
                DATEPART(HOUR, gp.PAR_TIME)            AS hh,
                CAST(gp.PAR_VALUE AS float)            AS par_value,
                ROW_NUMBER() OVER (
                    PARTITION BY RTRIM(p.pti), DATEPART(HOUR, gp.PAR_TIME)
                    ORDER BY gp.PAR_TIME DESC
                ) AS rn
            FROM [LOVATI].[dbo].[PTI] p
            JOIN [LOVATI].[dbo].[GacmPredictPTC] gp
              ON gp.pti = p.id
            WHERE 1=1
              {type_cond}
              AND RTRIM(p.pti) IN ({placeholders})
              AND gp.PAR_TIME >= ?
              AND gp.PAR_TIME < ?
              AND gp.PAR_VALUE IS NOT NULL
        )
        SELECT ptc_code, hh, par_value
        FROM src
        WHERE rn = 1
        ORDER BY ptc_code, hh;
    """

    templates: dict[str, dict[int, float]] = {}
    try:
        with pyodbc.connect(dsn_lovati, timeout=DB_CONNECT_TIMEOUT) as conn:
            cur = conn.cursor()
            cur.execute(sql, [*ptc_codes, day_start, day_end])
            for ptc_code, hh, par_value in cur.fetchall():
                ptc_code = (ptc_code or "").strip()
                if not ptc_code:
                    continue
                try:
                    hour = int(hh)
                    val = float(par_value)
                except Exception:
                    continue
                templates.setdefault(ptc_code, {})[hour] = val
    except Exception:
        return {}

    return templates


def _apply_gacm_template(rows: list[dict]) -> None:
    """
    Подставляет Gacm-P (gacm_p) по почасовым прогнозам из LOVATI:
      - для LOVATI объектов (src='lovati') берём прогнозы из GacmPredictPTC при p.typeObj = 0
      - для TERMOCOM объектов (src='termocom') берём прогнозы из GacmPredictPTC БЕЗ фильтра typeObj (по совпадению p.pti)

    Значение берём методом Zero-Order Hold (ZOH):
      - если на текущий час есть прогноз — берём его
      - если нет — берём последний прогноз <= текущего часа
      - если в текущем дне вообще нет — берём последнее значение предыдущего дня
    """
    if not rows:
        return

    # кеш: day -> {'lovati': templates, 'termocom': templates}
    template_cache: dict[date, dict[str, dict[str, dict[int, float]]]] = {}

    def _get_templates_for_day(day: date) -> dict[str, dict[str, dict[int, float]]]:
        if day in template_cache:
            return template_cache[day]

        termocom_ptc = sorted({
            (x.get("ptc") or "").strip()
            for x in rows
            if x.get("src") == "termocom" and x.get("ptc")
        })
        lovati_ptc = sorted({
            (x.get("ptc") or "").strip()
            for x in rows
            if x.get("src") == "lovati" and x.get("ptc")
        })

        tmpl_termocom = _load_gacm_hourly_template(termocom_ptc, day, typeobj_filter=None) if termocom_ptc else {}
        tmpl_lovati   = _load_gacm_hourly_template(lovati_ptc,   day, typeobj_filter="= 0")  if lovati_ptc else {}

        template_cache[day] = {"termocom": tmpl_termocom, "lovati": tmpl_lovati}
        return template_cache[day]

    def _zoh_value(src: str, ptc_code: str, day: date, hour: int) -> float | None:
        tmpl_by_src = _get_templates_for_day(day)
        hour_map = (tmpl_by_src.get(src) or {}).get(ptc_code) or {}

        if hour in hour_map:
            return hour_map[hour]

        prev_hours = [h for h in hour_map.keys() if isinstance(h, int) and h <= hour]
        if prev_hours:
            return hour_map.get(max(prev_hours))

        prev_day = day - timedelta(days=1)
        tmpl_prev_by_src = _get_templates_for_day(prev_day)
        hour_map_prev = (tmpl_prev_by_src.get(src) or {}).get(ptc_code) or {}
        if hour_map_prev:
            return hour_map_prev.get(max(hour_map_prev.keys()))

        return None

    for r in rows:
        ptc_code = (r.get("ptc") or "").strip()
        time_iso = r.get("time_iso")
        src = (r.get("src") or "").strip()  # 'lovati' или 'termocom'
        if not ptc_code or not time_iso or src not in ("lovati", "termocom"):
            continue

        try:
            base_dt = datetime.fromisoformat(str(time_iso))
        except Exception:
            continue

        if timezone.is_naive(base_dt):
            base_dt = base_dt.replace(tzinfo=TZ_CHISINAU)
        else:
            base_dt = base_dt.astimezone(TZ_CHISINAU)

        day = base_dt.date()
        hour = base_dt.hour

        val = _zoh_value(src, ptc_code, day, hour)
        if val is None:
            continue

        try:
            r["gacm_p"] = round(float(val), 2)
        except Exception:
            continue



def _to_int_or_none(x):
    try:
        return int(str(x).strip())
    except (TypeError, ValueError):
        return None


_id_re = re.compile(r"^(?=.*[A-Za-z])[A-Za-z0-9]+$")


def _is_valid_lovati_id(v) -> bool:
    """Валидный id_lovati: латиница+цифры, есть хотя бы одна буква; не пусто и не '0'."""
    s = "" if v is None else str(v).strip()
    return bool(s) and s != "0" and bool(_id_re.fullmatch(s))


def _url_1111_id_lovati(v) -> str | None:
    """
    РАНЬШЕ: строило ссылку на старый Streamlit по id_lovati.
    СЕЙЧАС: оставляем как есть (для совместимости), но в новой логике LOVATI используем _chart_url(..) в _id_only().
    """
    v = "" if v is None else str(v).strip()
    return f"http://10.1.1.248:1111/?id_lovati={v}" if _is_valid_lovati_id(v) else None


# >>> карта параметров Monitoring → имя параметра в LR Chart
_CHART_PARAM_MAP = {
    # потоки/мощность
    "q": "Q1",
    "q1": "Q1",
    "g1": "G1",
    "g2": "G2",
    "dg": "DG",
    "dt": "DT",
    "dg_pct": "DG",
    # температуры
    "t1": "T1",
    "t2": "T2",
    "t3": "T3",
    "t31": "T31",
    "t32": "T32",
    "t41": "T41",
    "t42": "T42",
    "t43": "T43",
    "t44": "T44",
    # acm/прочее
    "tacm": "TACM",
    "gacm": "GACM",
    "gacm_p": "GACM",
    "gadaos": "GADAOS",
    "sursa": "SURSA",
    "pompa": "POMPA",
    "pompa2": "POMPA2",
    "pompa3": "POMPA3",

}

# PTC, у которых значение Gacm нужно брать из G1 "A"-объекта
# 5019  <- G1(5019A)
# 4046  <- G1(4046A)
PTC_GACM_FROM_A = {
    "5019": "5019A",
    "4046": "4046A",
}


def _chart_url(ptc: str, param: str, start: str | None = None, end: str | None = None, agg: str | None = None) -> str:
    """
    Универсальный генератор ссылки на НОВЫЙ график LR Chart.
    Пример: /charts/chart/?pti=3107&param=T1
    """
    base = reverse("charts:chart_page")
    p = _CHART_PARAM_MAP.get((param or "").lower(), (param or "").upper())
    q = {"pti": str(ptc), "param": p}
    if start:
        q["from"] = start
    if end:
        q["to"] = end
    if agg:
        q["agg"] = agg
    return f"{base}?{urlencode(q)}"


def _url_1111_param(obiect: str, param: str) -> str:
    """
    Для объектов TERMOCOM5 строим ссылку на НОВУЮ страницу графика:
        /tc-charts/chart/?pti=1025&param=g1
    где:
      - pti  = номер PTC без префикса 'PT_'
      - param = имя параметра ('g1','t1','dt',...)
    """
    ptc = str(obiect or "").strip()
    if ptc.upper().startswith("PT_"):
        ptc = ptc[3:]

    base = reverse("termocom_charts:chart_page")  # /tc-charts/chart/
    q = {"pti": str(ptc), "param": str(param).lower()}
    return f"{base}?{urlencode(q)}"


def _quote_ident_sqlsrv(name: str) -> str:
    """Экранируем идентификатор MS SQL в []."""
    if not name:
        return name
    return "[" + name.replace("]", "]]") + "]"


def _get_ids_existing_columns(conn) -> set[str]:
    """Набор имён колонок таблицы IDS (в нижнем регистре)."""
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT LOWER(COLUMN_NAME) AS cn
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = 'IDS'
            """
        )
        return {str(r.cn).strip() for r in cur.fetchall()}
    except Exception:
        return set()


def _collect_ids_urls_by_pid(conn) -> dict[int, dict[str, str]]:
    """
    Читаем из IDS ID-значения для графиков по ключу PID (PTI.id).
    Возвращаем: { PID: {canon_param: id_lovati_str, ...}, ... }

    canon_param — наши канонические ключи:
      'q1','g1','g2','dg','dt','t1','t2','t3','t31','t32','t41','t42','t43','t44',
      'tacm','gacm','gadaos','sursa','pompa','pompa2'
    """
    candidates = {
        "q1": ["q1", "Q1"],
        "g1": ["G1", "g1"],
        "g2": ["G2", "g2"],
        "dg": ["dG", "DG", "dg"],
        "dt": ["dt", "DT"],
        "t1": ["T1", "t1"],
        "t2": ["T2", "t2"],
        "t3": ["T3", "t3"],
        "t31": ["T31", "t31"],
        "t32": ["T32", "t32"],
        "t41": ["T41", "t41"],
        "t42": ["T42", "t42"],
        "t43": ["T43", "t43"],
        "t44": ["T44", "t44"],
        "tacm": ["Tacm", "tacm"],
        "gacm": ["Gacm", "gacm"],
        "gadaos": ["Gadaos", "gadaos"],
        "sursa": ["sursa", "Sursa", "220V"],
        "pompa": ["pompa", "Pompa"],
        "pompa2": ["pompa2", "Pompa2"],
        "pompa3": ["pompa3", "Pompa3"],
    }

    existing = _get_ids_existing_columns(conn)

    col_for_key: dict[str, str] = {}
    for key, opts in candidates.items():
        found = None
        for c in opts:
            if c.lower() in existing:
                found = c
                break
        if found:
            col_for_key[key] = found

    if not col_for_key:
        return {}

    def _case_valid(col: str, require_d82_prefix: bool = False) -> str:
        q = _quote_ident_sqlsrv(col)
        extra_cond = ""
        if require_d82_prefix:
            # Для глубоких температур (T31–T44) допускаем только ID, начинающиеся на 'D82'
            extra_cond = f"WHEN LEFT(LTRIM(RTRIM({q})), 3) <> 'D82' THEN NULL "
        return (
            "CASE "
            f"WHEN {q} IS NULL THEN NULL "
            f"WHEN LTRIM(RTRIM({q})) IN ('','0') THEN NULL "
            f"WHEN LTRIM(RTRIM({q})) NOT LIKE '%[A-Za-z]%' THEN NULL "
            f"{extra_cond}"
            f"ELSE LTRIM(RTRIM({q})) "
            "END"
        )

    select_cols = ", ".join(
        f"{_case_valid(col, key in ('t31', 't32', 't41', 't42', 't43', 't44'))} AS [{key}]"
        for key, col in col_for_key.items()
    )
    sql = f"SELECT PTI AS PID_RAW, {select_cols} FROM IDS WHERE PTI IS NOT NULL"

    out: dict[int, dict[str, str]] = {}
    cur = conn.cursor()
    try:
        cur.execute(sql)
        for r in cur.fetchall():
            pid_raw = getattr(r, "PID_RAW", None)
            try:
                pid = int(pid_raw) if pid_raw is not None else None
            except Exception:
                pid = None
            if pid is None:
                continue

            m: dict[str, str] = {}
            for key in col_for_key.keys():
                val = getattr(r, key, None)
                sval = "" if val is None else str(val).strip()
                if _is_valid_lovati_id(sval):
                    m[key] = sval
            if m:
                out[pid] = m
    except Exception:
        return {}
    return out


# --- форматирование даты для фронтенда ---
def _fmt_frontend_dt(dt):
    """
    Вернёт строку в формате DD-MM-YY HH:MM или '' (локальная таймзона).
    Поддерживает naive/aware datetime.
    """
    try:
        if dt is None:
            return ""
        # Если datetime "наивный", считаем его временем Кишинёва (TZ_CHISINAU).
        if timezone.is_naive(dt):
            dt = dt.replace(tzinfo=TZ_CHISINAU)
        else:
            # Для aware-datetime переводим в часовой пояс Кишинёва,
            # чтобы корректно учесть зимнее/летнее время.
            dt = dt.astimezone(TZ_CHISINAU)
        return dt.strftime("%d-%m-%y %H:%M")
    except Exception:
        return ""


# ───────── core fetchers ─────────
def _fetch_termocom_rows():
    """
    TERMOCOM5: UNITS.UNIT_NAME вида PT_####/#####, нормализованные dict.
    Для графиков используем param_rokura (у нас нет PTI.id из LOVATI).
    """
    pompa_map = {
        "2009": [2],
        "2055": [2, 3],
        "2056": [2],
        "2057": [2],
        "2201": [2],
        "2202": [2, 3],
        "2209": [1],
        "2216": [2],
        "3012": [2],
        "3125": [1, 2, 3],
        "4009": [2],
        "4012": [2],
        "4014": [2],
        "4016": [2],
        "4019": [2],
        "4021": [2],
        "4025": [2],
        "4027": [2],
        "4037": [2],
        "4040": [2],
        "4041": [2],
        "4050": [2],
        "4054": [2],
        "4058": [2],
        "4063": [2],
        "4065": [2],
        "4066": [2],
        "4068": [2],
        "4077": [2],
        "5002": [2],
        "5003": [2],
        "5008": [2],
        "5009": [2],
        "5014": [2],
        "5019": [2],
        "5047": [2],
        "5057": [2],
        "5058": [2],
        "5075": [2],
    }

    # --- LOVATI: адреса по PTC ---
    dsn_lavati = _dsn(settings.LOVATI_SERVER)
    with pyodbc.connect(dsn_lavati, timeout=DB_CONNECT_TIMEOUT) as conn_l:
        cur_l = conn_l.cursor()

        # Прогноз Gacm-P по PTC (оставлено как было)
        cur_l.execute(
            """
            SELECT PTI, AVG(PAR_VALUE) AS AVG_PAR
            FROM GacmPredictPTC
            GROUP BY PTI
            """
        )
        _ = {int(r.PTI): float(r.AVG_PAR) for r in cur_l.fetchall()}

        # Адреса по PTC — теперь из таблицы [LOVATI].[dbo].[PTC_adrese]
        cur_l.execute(
            """
            SELECT
                RTRIM(PTC)    AS PTC,
                RTRIM(adresa) AS adresa
            FROM [LOVATI].[dbo].[PTC_adrese]
            WHERE LEN(RTRIM(PTC)) = 4
              AND LEFT(RTRIM(PTC), 1) IN ('1','2','3','4','5')
            """
        )
        address_map = {str(r.PTC).strip(): r.adresa for r in cur_l.fetchall()}

    # --- TERMOCOM5 ---
    dsn_termo = _dsn(settings.SQL_SERVER)
    with pyodbc.connect(dsn_termo, timeout=DB_CONNECT_TIMEOUT) as conn_t:
        cur_t = conn_t.cursor()
        cur_t.execute(
            """
            SELECT u.UNIT_ID,
                   u.UNIT_NAME,
                   mc.MC_T1_VALUE_INSTANT,
                   mc.MC_T2_VALUE_INSTANT,
                   mc.MC_G1_VALUE_INSTANT,
                   mc.MC_G2_VALUE_INSTANT,
                   mc.MC_POWER1_VALUE_INSTANT,
                   mc.MC_CINAVH_VALUE_INSTANT,
                   mc.MC_DTIME_VALUE_INSTANT,
                   mc.MC_DT_VALUE,
                   dcx.DCX_TR03_VALUE_INSTANT,
                   dcx.DCX_AI08_VALUE,
                   dcx.DCX_AI01_VALUE,
                   dcx.DCX_AI02_VALUE,
                   dcx.DCX_AI03_VALUE,
                   dcx.DCX_DTIME_VALUE_INSTANT,
                   dcx.DCX_CNT3_VALUE_INSTANT,
                   dcx.DCX_CNT4_VALUE_INSTANT,
                   comp.PT_MC_GINB_VALUE_INSTANT,
                   dcx.DCX_TR01_VALUE AS T31,
                   dcx.DCX_TR02_VALUE AS T32,
                   dcx.DCX_TR07_VALUE AS T41,
                   dcx.DCX_TR05_VALUE AS T42,
                   dcx.DCX_TR04_VALUE AS T43,
                   dcx.DCX_TR02_VALUE AS T44,
                   t3u.UNIT_LCS_VALUE
            FROM UNITS u
                     LEFT JOIN MULTICAL_CURRENT_DATA mc ON u.UNIT_ID = mc.UNIT_ID
                     LEFT JOIN DCX7600_CURRENT_DATA dcx ON u.UNIT_ID = dcx.UNIT_ID
                     LEFT JOIN PT_MC_COMPUTED_DATA comp ON u.UNIT_ID = comp.UNIT_ID
                     LEFT JOIN TERMOCOM3_UNIT t3u ON u.UNIT_ID = t3u.UNIT_ID
            WHERE u.UNIT_ENABLED = 1
              AND u.UNIT_NAME LIKE 'PT_%'
              AND (LEN(REPLACE(RTRIM(UNIT_NAME), 'PT_', '')) IN (4, 5))
            ORDER BY mc.MC_DTIME_VALUE_INSTANT DESC
            """
        )

        # 1) Считываем все строки разом
        rows = cur_t.fetchall()

        # 2) Собираем карту: базовый PTC -> G1 из "A"-объекта (5019A, 4046A)
        g1_from_A: dict[str, float] = {}
        for row in rows:
            ptc_full = row.UNIT_NAME.replace("PT_", "").strip()  # '5019', '5019A', ...
            if ptc_full.endswith("A"):
                base = ptc_full[:-1]  # '5019A' -> '5019'
                # проверяем, что это именно та пара, которая у нас описана в PTC_GACM_FROM_A
                if PTC_GACM_FROM_A.get(base) == ptc_full:
                    g1_from_A[base] = float(row.MC_G1_VALUE_INSTANT or 0.0)

        out = []

        # 3) Строим итоговые строки для таблицы
        for row in rows:
            ptc = row.UNIT_NAME.replace("PT_", "").strip()

            # 3a) сами объекты 5019A и 4046A в таблицу НЕ выводим
            if ptc in PTC_GACM_FROM_A.values():
                continue

            unit_id = int(row.UNIT_ID) if row.UNIT_ID is not None else None

            g1 = row.MC_G1_VALUE_INSTANT or 0
            g2 = row.MC_G2_VALUE_INSTANT or 0
            dg = g1 - g2
            # Δ% считаем только если оба значения реально есть и > 0
            dg_pct = None
            if g1 and g2:
                dg_pct = round(((g1 - g2) / g1) * 100.0, 1)

            v220_raw = row.DCX_AI08_VALUE or 0
            v220_on = v220_raw >= 12.5

            pompa_vals = None
            if ptc in pompa_map:
                pompa_vals = []
                for num in pompa_map[ptc]:
                    if num == 1:
                        pompa_vals.append(row.DCX_AI01_VALUE or 0)
                    elif num == 2:
                        pompa_vals.append(row.DCX_AI02_VALUE or 0)
                    elif num == 3:
                        pompa_vals.append(row.DCX_AI03_VALUE or 0)

            obiect = f"PT_{ptc}"

            # форматируем время (ВАЖНО: если в TERMOCOM нет времени, используем текущее локальное,
            # иначе _apply_gacm_template пропустит строку из-за пустого time_iso)
            if row.MC_DTIME_VALUE_INSTANT:
                termocom_time = _fmt_frontend_dt(row.MC_DTIME_VALUE_INSTANT)
                termocom_time_iso = row.MC_DTIME_VALUE_INSTANT.isoformat(timespec="minutes")
            else:
                now_local = timezone.now()
                if timezone.is_naive(now_local):
                    now_local = now_local.replace(tzinfo=TZ_CHISINAU)
                else:
                    now_local = now_local.astimezone(TZ_CHISINAU)
                termocom_time = _fmt_frontend_dt(now_local)
                termocom_time_iso = now_local.isoformat(timespec="minutes")

            # T31–T44: скрываем нули (0 → None, в таблице будет пусто)
            def _hide_zero_t(v, ndigits=1):
                if v is None:
                    return None
                try:
                    val = round(v, ndigits)
                except Exception:
                    return None
                return None if val == 0 else val

            t31_val = _hide_zero_t(row.T31, 1)
            t32_val = _hide_zero_t(row.T32, 1)
            t41_val = _hide_zero_t(row.T41, 1)
            t42_val = _hide_zero_t(row.T42, 1)
            t43_val = _hide_zero_t(row.T43, 1)
            t44_val = _hide_zero_t(row.T44, 1)

            # 3b) Gacm: для 5019 и 4046 берём G1 из 5019A/4046A, для остальных — как было
            if ptc in g1_from_A:
                gacm_value = g1_from_A[ptc]
            else:
                gacm_value = row.MC_CINAVH_VALUE_INSTANT or 0

            out.append(
                {
                    "src": "termocom",
                    "ptc": ptc,
                    "address": address_map.get(ptc, ""),
                    "t1": round(row.MC_T1_VALUE_INSTANT or 0, 1),
                    "id_t1": _url_1111_param(obiect, "t1"),
                    "t2": round(row.MC_T2_VALUE_INSTANT or 0, 1),
                    "id_t2": _url_1111_param(obiect, "t2"),
                    "t3": round(row.DCX_CNT3_VALUE_INSTANT or 0),
                    "t4": round(row.DCX_CNT4_VALUE_INSTANT or 0),
                    "t31": t31_val,
                    "id_t31": _url_1111_param(obiect, "t31"),
                    "t32": t32_val,
                    "id_t32": _url_1111_param(obiect, "t32"),
                    "t41": t41_val,
                    "id_t41": _url_1111_param(obiect, "t41"),
                    "t42": t42_val,
                    "id_t42": _url_1111_param(obiect, "t42"),
                    "t43": t43_val,
                    "id_t43": _url_1111_param(obiect, "t43"),
                    "t44": t44_val,
                    "id_t44": _url_1111_param(obiect, "t44"),
                    "g1": round(g1, 2),
                    "id_g1": _url_1111_param(obiect, "g1"),
                    "g2": round(g2, 2),
                    "id_g2": _url_1111_param(obiect, "g2"),
                    "q1": round(row.MC_POWER1_VALUE_INSTANT or 0, 2),
                    "id_q1": _url_1111_param(obiect, "q"),
                    "dg": round(dg, 2),
                    "id_dg": _url_1111_param(obiect, "dg"),
                    "dt": round(row.MC_DT_VALUE or 0, 2),
                    "id_dt": _url_1111_param(obiect, "dt"),
                    "dg_pct": "" if dg_pct is None else dg_pct,
                    "id_dg_pct": _url_1111_param(obiect, "dg_pct"),

                    # ВАЖНО: здесь теперь используем gacm_value
                    "gacm": round(gacm_value, 2),
                    "id_gacm": _url_1111_param(obiect, "gacm"),
                    "gacm_p": "",
                    "tacm": round(row.DCX_TR03_VALUE_INSTANT or 0, 1),
                    "id_tacm": _url_1111_param(obiect, "tacm"),
                    "g_adaos": round((getattr(row, "PT_MC_GINB_VALUE_INSTANT", 0) or 0), 2),
                    "id_g_adaos": _url_1111_param(obiect, "gadaos"),
                    "sursa": v220_on,
                    "id_sursa": _url_1111_param(obiect, "sursa"),
                    "pompa": pompa_vals,
                    "pompa_nums": pompa_map.get(ptc, []),
                    "id_pompa1": _url_1111_param(obiect, "pompa"),
                    "id_pompa2": _url_1111_param(obiect, "pompa2"),
                    "id_pompa3": _url_1111_param(obiect, "pompa3"),
                    "lcs": round((row.UNIT_LCS_VALUE or 0) * 100, 2),
                    "time": termocom_time,
                    "time_iso": termocom_time_iso,
                }
            )

        return out


def _fetch_lovati_rows():
    """
    LOVATI: вытягиваем текущие значения из PTI и строим ссылки на графики.
    ДОБАВЛЕНО:
      - берём дату/время из p.dt1; если NULL — в поле 'time' кладём ''.
      - для T31–T44 и помп (pompa/pompa2) получаем реальные значения через HTTP+XML из приборов LR
        (последнее значение за последний 1 час) при наличии валидного id_lovati в таблице IDS.
      - запросы к LR выполняем параллельно для ускорения.
      - помпа1 ВСЕГДА берётся из PTI.Pompa (как минимум), даже если нет IDS/графиков.
      - Gacm-P берём из таблицы GacmPredictPTC с интерполяцией по ближайшим точкам.
    """
    dsn_lovati = _dsn(settings.LOVATI_SERVER)
    with pyodbc.connect(dsn_lovati, timeout=DB_CONNECT_TIMEOUT) as conn:
        cur = conn.cursor()

        # --- базовые данные PTI ---
        cur.execute(
            r"""
            SELECT p.id                                     AS PID, -- ключ для связи с IDS
                   RTRIM(p.pti)                             AS PTC,
                   RTRIM(p.adres_unicode)                   AS Adresa,
                   p.IPs                                    AS IPs,
                   p.dt1                                    AS dt1, -- дата/время последнего показания
                   ROUND(CAST(RTRIM(p.q1) AS float), 2)     AS q1,
                   ROUND(CAST(RTRIM(p.g1) AS float), 2)     AS G1,
                   ROUND(CAST(RTRIM(p.g2) AS float), 2)     AS G2,
                   ROUND(CAST(RTRIM(p.t1) AS float), 2)     AS T1,
                   ROUND(CAST(RTRIM(p.t2) AS float), 2)     AS T2,
                   ROUND(CAST(RTRIM(p.tacm) AS float), 2)   AS Tacm,
                   ROUND(CAST(RTRIM(p.gacm) AS float), 2)   AS Gacm,
                   ROUND(CAST(RTRIM(p.gadaos) AS float), 2) AS Gadaos,
                   ROUND(CAST(RTRIM(p.sursa) AS float), 2)  AS V220,
                   ROUND(CAST(RTRIM(p.pompa)  AS float), 2) AS Pompa,
                   ROUND(CAST(RTRIM(p.pompa2) AS float), 2) AS Pompa2,
                   ROUND(CAST(RTRIM(p.pompa3) AS float), 2) AS Pompa3

            FROM PTI p
            WHERE p.typeObj = 0
              AND LEN(RTRIM(p.pti)) = 4
              AND (LEFT(RTRIM(p.pti), 1) IN ('1','2','3','4','5'))
            ORDER BY p.pti
            """
        )
        rows = cur.fetchall()

        # IDS: id_lovati для ВСЕХ поддержанных параметров (по PID)
        ids_urls_by_pid = _collect_ids_urls_by_pid(conn)

        # --- прогноз Gacm-P для ВСЕХ доступных дат ---
        forecasts_by_ptc: dict[str, list[tuple[datetime, float]]] = defaultdict(list)
        cur.execute(
            """
            SELECT
                RTRIM(p.pti) AS ptc_code,
                gp.PAR_TIME  AS par_time,
                gp.PAR_VALUE AS par_value
            FROM GacmPredictPTI gp
            JOIN PTI p ON gp.pti = p.id
            WHERE p.typeObj = 0
              AND LEN(RTRIM(p.pti)) = 4
              AND LEFT(RTRIM(p.pti), 1) IN ('1','2','3','4','5')
            """
        )
        for fr in cur.fetchall():
            ptc_code = str(fr.ptc_code).strip()
            par_time = fr.par_time
            par_value = _to_float(fr.par_value)
            forecasts_by_ptc[ptc_code].append((par_time, par_value))

    # сортируем прогнозы по времени
    for ptc_code in forecasts_by_ptc:
        forecasts_by_ptc[ptc_code].sort(key=lambda x: x[0])

    # --- текущее время (локальное) для интерполяции и окна LR ---
    now_local = timezone.now()
    if timezone.is_naive(now_local):
        now_local = now_local.replace(tzinfo=TZ_CHISINAU)
    else:
        now_local = now_local.astimezone(TZ_CHISINAU)

    def _interp_gacm_p_for_ptc(ptc_code: str, target_dt_naive: datetime) -> float | None:
        """
        Линейная интерполяция Gacm-P для заданного времени target_dt_naive:
        - если точек нет — None;
        - если одна точка — эта точка;
        - если вне диапазона — берём ближайшую граничную;
        - иначе — линейная интерполяция между соседними часами.
        """
        pts = forecasts_by_ptc.get(str(ptc_code))
        if not pts:
            return None
        if len(pts) == 1:
            return pts[0][1]

        if target_dt_naive <= pts[0][0]:
            return pts[0][1]
        if target_dt_naive >= pts[-1][0]:
            return pts[-1][1]

        for i in range(1, len(pts)):
            t0, v0 = pts[i - 1]
            t1, v1 = pts[i]
            if t0 <= target_dt_naive <= t1:
                if t1 == t0:
                    return v1
                total = (t1 - t0).total_seconds()
                alpha = (target_dt_naive - t0).total_seconds() / total
                return v0 + (v1 - v0) * alpha
        return pts[-1][1]

    # --- окно 1 час для LR ---
    stop_epoch = to_epoch_seconds(now_local)
    start_epoch = to_epoch_seconds(now_local - timedelta(hours=1))

    # --- собираем пары (ips, id_lovati) для LR ---
    needed_pairs: set[tuple[int, str]] = set()
    parsed_rows = []
    for r in rows:
        pid = int(getattr(r, "PID"))
        ptc = str(r.PTC).strip()
        address = str(r.Adresa or "").strip()
        ips = _to_int_or_none(getattr(r, "IPs", None))
        ids_map = ids_urls_by_pid.get(pid, {})

        parsed_rows.append((r, pid, ptc, address, ips, ids_map))

        if ips:
            for key in ("t31", "t32", "t41", "t42", "t43", "t44", "pompa", "pompa2"):
                param_id = ids_map.get(key)
                if param_id:
                    needed_pairs.add((ips, param_id))

    # --- параллельные запросы к LR ---
    lr_cache: dict[tuple[int, str], float | None] = {}

    def _fetch_pair(ips: int, param_id: str) -> float | None:
        if not ips or not param_id:
            return None
        now_ts = timezone.now().timestamp()
        cache_key = (ips, str(param_id))

        cached = LR_MEMORY_CACHE.get(cache_key)
        if cached is not None:
            val, ts = cached
            if now_ts - ts < LR_CACHE_TTL:
                return val

        try:
            raw_xml = fetch_xml(ips, param_id, start_epoch, stop_epoch)
            series = parse_series(raw_xml)
            value = series[-1][1] if series else None
        except Exception:
            value = None

        LR_MEMORY_CACHE[cache_key] = (value, now_ts)
        return value

    if needed_pairs:
        max_workers = min(8, len(needed_pairs))
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_key = {
                executor.submit(_fetch_pair, ips, param_id): (ips, param_id)
                for (ips, param_id) in needed_pairs
            }
            for fut in as_completed(future_to_key):
                key = future_to_key[fut]
                try:
                    lr_cache[key] = fut.result()
                except Exception:
                    lr_cache[key] = None

    def _val_or_empty(value, id_url=None):
        return "" if value is None else value

    def _id_only(ptc: str, ids_map: dict, key: str) -> str | None:
        if not ids_map.get(key):
            return None
        return _chart_url(ptc, key)

    out = []
    for r, pid, ptc, address, ips, ids_map in parsed_rows:
        g1 = _to_float(r.G1)
        g2 = _to_float(r.G2)
        dg = round(g1 - g2, 2)
        # Δ% считаем только если оба значения реально есть и > 0
        dg_pct = None
        if g1 and g2:
            dg_pct = round(((g1 - g2) / g1) * 100.0, 1)

        # ✅ прогноз Gacm-P: интерполяция строго на +1 час от dt1 (времени объекта)
        dt1 = getattr(r, "dt1", None)
        gacm_p_value = None

        if dt1:
            # dt1 из SQL часто naive -> считаем, что это локальное время (Кишинёв)
            if timezone.is_naive(dt1):
                dt1_local = dt1.replace(tzinfo=TZ_CHISINAU)
            else:
                dt1_local = dt1.astimezone(TZ_CHISINAU)

            target_dt = dt1_local  # ✅ прогноз на текущее время прибора
            target_dt_naive = target_dt.replace(tzinfo=None)
            gacm_p_value = _interp_gacm_p_for_ptc(ptc, target_dt_naive)

        def _get_v(key: str) -> float | None:
            param_id = ids_map.get(key)
            if not ips or not param_id:
                return None
            return lr_cache.get((ips, param_id))

        v_t31 = _get_v("t31")
        v_t32 = _get_v("t32")
        v_t41 = _get_v("t41")
        v_t42 = _get_v("t42")
        v_t43 = _get_v("t43")
        v_t44 = _get_v("t44")

        v_pompa1 = _get_v("pompa")
        v_pompa2 = _get_v("pompa2")

        def _hide_zero_round(v, ndigits: int = 1):
            if v is None:
                return None
            val = _roundf(v, ndigits)
            return None if val == 0 else val

        t31_val = _hide_zero_round(v_t31, 1)
        t32_val = _hide_zero_round(v_t32, 1)
        t41_val = _hide_zero_round(v_t41, 1)
        t42_val = _hide_zero_round(v_t42, 1)
        t43_val = _hide_zero_round(v_t43, 1)
        t44_val = _hide_zero_round(v_t44, 1)

        pompa_vals = None
        pompa_nums: list[int] = []

        # ✅ Если это “наши” LOVATI объекты — берём статусы насосов прямо из PTI: r.Pompa/r.Pompa2/r.Pompa3
        if ptc in LOVATI_PUMP01_PTC:
            raw_pumps_01: list[tuple[int, int]] = []

            def _to01(x):
                if x is None:
                    return None
                try:
                    v = int(float(x))
                except Exception:
                    return None
                return v if v in (0, 1) else None

            s1 = _to01(getattr(r, "Pompa", None))
            s2 = _to01(getattr(r, "Pompa2", None))
            s3 = _to01(getattr(r, "Pompa3", None))

            # ✅ pompa1 показываем, если есть 0/1 (ID в IDS не обязателен)
            if s1 is not None:
                raw_pumps_01.append((1, s1))

            # ✅ pompa2/pompa3 показываем ТОЛЬКО если в IDS есть реальный ID (не 0/NULL)
            if ids_map.get("pompa2") and s2 is not None:
                raw_pumps_01.append((2, s2))

            if ids_map.get("pompa3") and s3 is not None:
                raw_pumps_01.append((3, s3))

            if raw_pumps_01:
                pompa_nums = [num for (num, _v) in raw_pumps_01]
                pompa_vals = [_v for (_num, _v) in raw_pumps_01]


        else:

            # ❌ Для остальных LOVATI объектов помпы НЕ показываем вообще

            pompa_vals = None

            pompa_nums = []

        id_q1 = _id_only(ptc, ids_map, "q1")
        id_g1 = _id_only(ptc, ids_map, "g1")
        id_g2 = _id_only(ptc, ids_map, "g2")
        id_dg = _id_only(ptc, ids_map, "dg")
        id_dt = _id_only(ptc, ids_map, "dt")
        id_t1 = _id_only(ptc, ids_map, "t1")
        id_t2 = _id_only(ptc, ids_map, "t2")
        id_t3 = _id_only(ptc, ids_map, "t3")
        id_t31 = _id_only(ptc, ids_map, "t31")
        id_t32 = _id_only(ptc, ids_map, "t32")
        id_t41 = _id_only(ptc, ids_map, "t41")
        id_t42 = _id_only(ptc, ids_map, "t42")
        id_t43 = _id_only(ptc, ids_map, "t43")
        id_t44 = _id_only(ptc, ids_map, "t44")
        id_tacm = _id_only(ptc, ids_map, "tacm")
        id_gacm = _id_only(ptc, ids_map, "gacm")
        id_gadaos = _id_only(ptc, ids_map, "gadaos")
        id_sursa = _id_only(ptc, ids_map, "sursa")
        id_pompa1 = _id_only(ptc, ids_map, "pompa")

        dt1 = getattr(r, "dt1", None)
        if dt1:
            time_str = _fmt_frontend_dt(dt1) if hasattr(dt1, "strftime") else str(dt1).strip()[:16]
            time_iso = dt1.isoformat(timespec="minutes") if hasattr(dt1, "isoformat") else ""
        else:
            time_str = ""
            time_iso = ""

        sursa_flag = "" if not id_sursa else (_to_float(r.V220) >= 12.5)

        # --- фильтруем помпы: показываем только те, где реально есть id в IDS для pompa2/pompa3 ---
        allowed_nums: list[int] = []
        allowed_vals: list[int] | list[float] = []

        if isinstance(pompa_nums, list) and isinstance(pompa_vals, list):
            for num, val in zip(pompa_nums, pompa_vals):
                if num == 1:
                    # pompa1 всегда можно показывать (она "главная")
                    allowed_nums.append(num)
                    allowed_vals.append(val)
                elif num == 2:
                    # pompa2 показываем только если в IDS есть реальный ID (не 0/NULL)
                    if ids_map.get("pompa2"):
                        allowed_nums.append(num)
                        allowed_vals.append(val)
                elif num == 3:
                    # pompa3 показываем только если в IDS есть реальный ID (не 0/NULL)
                    if ids_map.get("pompa3"):
                        allowed_nums.append(num)
                        allowed_vals.append(val)

        # финальные значения для фронта
        pompa_nums = allowed_nums
        pompa_vals = allowed_vals if allowed_vals else None


        out.append(
            {
                "src": "lovati",
                "ptc": ptc,
                "address": address,
                "q1": _val_or_empty(_roundf(r.q1, 2), id_q1),
                "id_q1": id_q1,
                "g1": _val_or_empty(_roundf(g1, 2), id_g1),
                "id_g1": id_g1,
                "g2": _val_or_empty(_roundf(g2, 2), id_g2),
                "id_g2": id_g2,
                "dg": _val_or_empty(dg, id_dg),
                "id_dg": id_dg,
                "dg_pct": _val_or_empty(dg_pct),
                "dt": _val_or_empty(_roundf(_to_float(r.T1) - _to_float(r.T2), 2), id_dt),
                "id_dt": id_dt,
                "t1": _val_or_empty(_roundf(r.T1, 1), id_t1),
                "id_t1": id_t1,
                "t2": _val_or_empty(_roundf(r.T2, 1), id_t2),
                "id_t2": id_t2,
                "t3": _val_or_empty(0, id_t3),
                "id_t3": id_t3,
                "t4": "",
                "t31": _val_or_empty(t31_val, id_t31),
                "id_t31": id_t31,
                "t32": _val_or_empty(t32_val, id_t32),
                "id_t32": id_t32,
                "t41": _val_or_empty(t41_val, id_t41),
                "id_t41": id_t41,
                "t42": _val_or_empty(t42_val, id_t42),
                "id_t42": id_t42,
                "t43": _val_or_empty(t43_val, id_t43),
                "id_t43": id_t43,
                "t44": _val_or_empty(t44_val, id_t44),
                "id_t44": id_t44,
                "gacm": _val_or_empty(_roundf(r.Gacm, 2), id_gacm),
                "id_gacm": id_gacm,
                "gacm_p": _disp_num(gacm_p_value, 2),
                "tacm": _val_or_empty(_roundf(r.Tacm, 1), id_tacm),
                "id_tacm": id_tacm,
                "g_adaos": _val_or_empty(_roundf(r.Gadaos, 2), id_gadaos),
                "id_g_adaos": id_gadaos,
                "sursa": sursa_flag,
                "id_sursa": id_sursa,


                "pompa": pompa_vals,
                "pompa_nums": pompa_nums,
                "id_pompa1": _chart_url(ptc, "pompa")  if ids_map.get("pompa")  else None,
                "id_pompa2": _chart_url(ptc, "pompa2") if ids_map.get("pompa2") else None,
                "id_pompa3": _chart_url(ptc, "pompa3") if ids_map.get("pompa3") else None,



                "lcs": 0.0 if not sursa_flag else 100.0,
                "time": time_str,
                "time_iso": time_iso,
            }
        )

    return out


def fetch_ptc_data():
    """TERMOCOM5 + LOVATI с de-dup по PTC. Приоритет у TERMOCOM5.
    ДОПОЛНИТЕЛЬНО: рассчитываем Gacm-P (прогноз) по шаблону суток из LOVATI.
    """
    termo_rows = _fetch_termocom_rows()
    termo_ptc = {row["ptc"] for row in termo_rows}

    lovati_rows = _fetch_lovati_rows()
    # LOVATI-строки с PTC, которые уже есть в TERMOCOM5, убираем
    lovati_rows = [r for r in lovati_rows if r.get("ptc") not in termo_ptc]

    # объединяем
    rows = termo_rows + lovati_rows

    # здесь аккуратно подставляем Gacm-P (игнорируя PAR_VALUE = 0 в шаблоне)
    _apply_gacm_template(rows)

    return rows


# ───────── Django views ─────────
def ptc_table(request):
    return render(request, "monitoring/ptc_table.html", {"can_edit": can_edit_from_request(request)})


def api_ptc_data(request):
    season = (request.GET.get("season") or "Iarna").strip()
    season = season if season in ("Iarna", "Vara", "Toate") else "Iarna"

    def _flag(name):
        val = (request.GET.get(name) or "").strip().lower()
        return val in ("1", "true", "on", "yes")

    # --- флаги ---
    t1_en = _flag("t1min_enabled")  # 1. T1 min
    t4_en = _flag("t4min_enabled")  # 2. T4 min
    dt_en = _flag("dtmin_enabled")  # 3. ΔT min
    tacm_en = _flag("tacm_enabled")  # 4. Tacm
    gacm_max_en = _flag("gacm_max_enabled")  # 5. Gacm max
    dgacm_en = _flag("dgacm_enabled")  # 6. ΔGacm max
    g1_min_en = _flag("g1_min_enabled")  # 7. G1 min
    dgp_en = _flag("dgp_enabled")  # 8. ΔG% max
    dg_flow_en = _flag("dg_flow_enabled")  # 9. ΔG max
    gadaos_en = _flag("gadaos_enabled")  # 10. Gadaos max
    dataora_en = _flag("dataora_enabled")  # 11. Ore fără date
    pompa_off_en = _flag("pompa_off_enabled")  # 12. Pompa OFF (roșu)
    sursa_off_en = _flag("sursa_off_enabled")  # 13. 220V OFF (roșu)

    any_filter_enabled = any(
        [
            t1_en,
            t4_en,
            dt_en,
            tacm_en,
            gacm_max_en,
            dgacm_en,
            g1_min_en,
            dgp_en,
            dg_flow_en,
            gadaos_en,
            dataora_en,
            pompa_off_en,
            sursa_off_en,
        ]
    )

    # --- пороги ---
    t1_thr = _to_float(request.GET.get("t1min_t1"), 50.0)
    g1_thr = _to_float(request.GET.get("t1min_g1"), 0.1)

    t4_thr = _to_float(request.GET.get("t4min_t4"), 30.0)

    dt_thr = _to_float(request.GET.get("dtmin_dt"), 5.0)
    t1_over = _to_float(request.GET.get("dtmin_t1_over"), 50.0)

    tacm_min = _to_float(request.GET.get("tacm_min"), 50.0)
    tacm_max = _to_float(request.GET.get("tacm_max"), 60.0)

    gacm_max_limit = _to_float(request.GET.get("gacm_max"), 10.0)

    dgacm_split = _to_float(request.GET.get("dgacm_split"), 5.0)  # порог переключения режимов сравнения
    dgacm_abs = _to_float(request.GET.get("dgacm_abs"), 1.0)  # ΔG ≥ 1.0
    dgacm_pct = _to_float(request.GET.get("dgacm_pct"), 20.0)  # ΔG% ≥ 20

    g1_min_limit = _to_float(request.GET.get("g1_min"), 0.5)
    dgp_limit = _to_float(request.GET.get("dgp_limit"), 2.5)
    dg_flow_limit = _to_float(request.GET.get("dg_flow_limit"), 1.0)
    gadaos_limit = _to_float(request.GET.get("gadaos_limit"), 0.1)
    try:
        dataora_limit = int(request.GET.get("dataora_limit") or 1)
    except (TypeError, ValueError):
        dataora_limit = 1

    data = fetch_ptc_data()

    # пометка строк, у которых есть комментарии
    comments_map = _load_json(COMMENTS_PATH)

    def _has_comments(ptc: str) -> bool:
        return bool(comments_map.get(str(ptc) or ""))

    for r in data:
        r["has_comment"] = _has_comments(r.get("ptc", ""))

    # --- исключения: берём активные на сегодня (tura из GET, по умолчанию "1") ---
    tura = str(request.GET.get("tura") or "1")
    active = _get_active_exclusions(today=timezone.localdate(), tura_cur=tura)


    if active:
        for r in data:
            ptc_code = str(r.get("ptc") or "")
            r["has_exclusion"] = bool(active.get(ptc_code))

    EX_SYNONYMS = {
        "t1": {"t1"},
        "t2": {"t2", "dt"},
        "t41": {"t41", "t4"},
        "t42": {"t42", "t4"},
        "t43": {"t43", "t4"},
        "t44": {"t44", "t4"},
        "tacm": {"tacm"},
        "gacm": {"gacm"},
        "dgacm": {"gacm_p", "dgacm_val", "dgacm"},
        "g1": {"g1"},
        "dg_pct": {"dg_pct"},
        "dg": {"dg"},
        "g_adaos": {"g_adaos"},
        "time": {"time"},
    }

    def _is_excluded(key: str, exset: set[str]) -> bool:
        syn = EX_SYNONYMS.get(key, {key})
        return any(s in exset for s in syn)

    def _has_measured_t(r: dict, key: str) -> bool:
        """
        Есть ли реальное значение для T4-датчика.
        Пустые ''/None считаем «нет измерения», даже если id_* есть.
        """
        idk = "id_" + key
        v = r.get(key)

        # если в таблице значение пустое/None (в том числе после скрытия нулей) – измерения нет
        if v in ("", None):
            return False

        # если есть id_*, он тоже должен быть не пустой
        if idk in r:
            return bool(r.get(idk))

        # иначе достаточно самого значения
        return True

    # --- вычисляем триггеры/подсветку ---
    for r in data:
        # 1. T1 min
        r["t1_trigger"] = (_to_float(r.get("t1")) <= t1_thr) and (_to_float(r.get("g1")) > g1_thr)

        # 2. T4 min
        if t4_en:
            r["t41_red"] = _has_measured_t(r, "t41") and (_to_float(r.get("t41")) <= t4_thr)
            r["t42_red"] = _has_measured_t(r, "t42") and (_to_float(r.get("t42")) <= t4_thr)
            r["t43_red"] = _has_measured_t(r, "t43") and (_to_float(r.get("t43")) <= t4_thr)
            r["t44_red"] = _has_measured_t(r, "t44") and (_to_float(r.get("t44")) <= t4_thr)
        else:
            r["t41_red"] = r["t42_red"] = r["t43_red"] = r["t44_red"] = False

        # 3. ΔT min (красим T2)
        if dt_en:
            r["t2_red"] = (_to_float(r.get("dt")) < dt_thr) and (_to_float(r.get("t1")) > t1_over)
        else:
            r["t2_red"] = False

        # 4. Tacm
        if tacm_en:
            r["tacm_red"] = (_to_float(r.get("tacm")) <= tacm_min) or (_to_float(r.get("tacm")) >= tacm_max)
        else:
            r["tacm_red"] = False

        # 5. Gacm max
        if gacm_max_en:
            r["gacm_red"] = _to_float(r.get("gacm")) >= gacm_max_limit
        else:
            r["gacm_red"] = False

        # 6. ΔGacm max (ACM) → красит Gacm-P
        if dgacm_en:
            v_gacm = _to_float(r.get("gacm"))
            v_gacm_p = _to_float(r.get("gacm_p"))
            d_gacm = v_gacm - v_gacm_p
            d_pct = (d_gacm / v_gacm_p) * 100.0 if v_gacm_p > 0 else 0.0
            if v_gacm >= dgacm_split:
                r["dgacm_red"] = d_gacm >= dgacm_abs
            else:
                r["dgacm_red"] = d_pct >= dgacm_pct
        else:
            r["dgacm_red"] = False

        # 7. G1 min
        if g1_min_en:
            r["g1_red"] = _to_float(r.get("g1")) <= g1_min_limit
        else:
            r["g1_red"] = False

        # 8. ΔG% max
        if dgp_en:
            g1_val = _to_float(r.get("g1"))
            g2_val = _to_float(r.get("g2"))
            dgp = ((g1_val - g2_val) / g1_val * 100.0) if g1_val else 0
            r["dgp_red"] = dgp > dgp_limit
        else:
            r["dgp_red"] = False

        # 9. ΔG max
        if dg_flow_en:
            g1_val = _to_float(r.get("g1"))
            g2_val = _to_float(r.get("g2"))
            r["dg_flow_red"] = (g1_val - g2_val) > dg_flow_limit
        else:
            r["dg_flow_red"] = False

        # 10. Gadaos max
        if gadaos_en:
            r["gadaos_red"] = _to_float(r.get("gadaos")) > gadaos_limit
        else:
            r["gadaos_red"] = False

        # 11. Ore fără date
        if dataora_en:
            ts_raw = r.get("time_iso") or r.get("time") or ""
            parsed = None

            try:
                parsed = datetime.fromisoformat(str(ts_raw))
            except Exception:
                parsed = None

            if parsed is None:
                try:
                    parsed = datetime.strptime(str(ts_raw), "%d-%m-%y %H:%M")
                except Exception:
                    parsed = None

            if parsed is not None:
                # Если временная зона не указана, считаем dt локальным
                # временем Кишинёва; иначе переводим в этот часовой пояс.
                if parsed.tzinfo is None:
                    parsed = parsed.replace(tzinfo=TZ_CHISINAU)
                else:
                    parsed = parsed.astimezone(TZ_CHISINAU)

                # Текущее время также берётся в зоне Кишинёва.
                now = datetime.now(TZ_CHISINAU)
                delta_h = (now - parsed).total_seconds() / 3600
                r["dataora_red"] = delta_h > dataora_limit
            else:
                r["dataora_red"] = False
        else:
            r["dataora_red"] = False

            # 12. Pompa OFF (красная помпа)
        if pompa_off_en:
            vals = r.get("pompa") or []
            if isinstance(vals, list) and vals:

                floats = [_to_float(v, None) for v in vals]

                # LOVATI: цифровые статусы 0/1 (0=ON, 1=OFF)
                is_digital_01 = True
                for v in floats:
                    if v is None:
                        continue
                    if v not in (0.0, 1.0):
                        is_digital_01 = False
                        break

                if is_digital_01:
                    # ТРЕБОВАНИЕ: все LOVATI с 1=OFF → Pompa oprită
                    r["pompa_off"] = any(v == 1.0 for v in floats if v is not None)
                else:
                    # TERMOCOM5 / аналог: старая логика НЕ МЕНЯЕТСЯ
                    lcs = _to_float(r.get("lcs"), 0.0)
                    LCS_NORM = 30.0

                    if lcs >= LCS_NORM:
                        r["pompa_off"] = any(_to_float(v, 0.0) > 200.0 for v in vals)
                    else:
                        r["pompa_off"] = False
            else:
                r["pompa_off"] = False
        else:
            r["pompa_off"] = False

            # 13. 220V OFF (красный кружок)
        if sursa_off_en:
            has_220_obj = bool(r.get("id_sursa"))
            is_on = bool(r.get("sursa"))
            # объект есть, но 220V нет → OFF
            r["sursa_off"] = has_220_obj and not is_on
        else:
            r["sursa_off"] = False

    # --- вычисляем "красноту" ДО применения исключений ---
    before_map: dict[str, set[str]] = {}
    for r in data:
        ptc = str(r.get("ptc", ""))
        red_keys: set[str] = set()
        if r.get("t1_trigger"):
            red_keys.add("t1")
        if r.get("t2_red"):
            red_keys.add("t2")
        if r.get("t41_red"):
            red_keys.add("t41")
        if r.get("t42_red"):
            red_keys.add("t42")
        if r.get("t43_red"):
            red_keys.add("t43")
        if r.get("t44_red"):
            red_keys.add("t44")
        if r.get("tacm_red"):
            red_keys.add("tacm")
        if r.get("gacm_red"):
            red_keys.add("gacm")
        if r.get("dgacm_red"):
            red_keys.add("dgacm")
        if r.get("g1_red"):
            red_keys.add("g1")
        if r.get("dgp_red"):
            red_keys.add("dg_pct")
        if r.get("dg_flow_red"):
            red_keys.add("dg")
        if r.get("gadaos_red"):
            red_keys.add("g_adaos")
        if r.get("dataora_red"):
            red_keys.add("time")
        if r.get("pompa_off"):
            red_keys.add("pompa")
        if r.get("sursa_off"):
            red_keys.add("sursa")
        before_map[ptc] = red_keys

    # --- применяем исключения ---
    if active:
        for r in data:
            ptc = str(r.get("ptc", ""))
            exset = active.get(ptc, set())
            EX = exset.__contains__
            ex_all = EX("__all__")

            r["excluded_all"] = bool(ex_all)

            if ex_all:
                r["t1_trigger"] = False
                r["t41_red"] = r["t42_red"] = r["t43_red"] = r["t44_red"] = False
                r["t2_red"] = False
                r["tacm_red"] = False
                r["gacm_red"] = False
                r["dgacm_red"] = False
                r["g1_red"] = False
                r["dgp_red"] = False
                r["dg_flow_red"] = False
                r["gadaos_red"] = False
                r["dataora_red"] = False
                r["pompa_off"] = False
                r["sursa_off"] = False
                continue

            if EX("t1"):
                r["t1_trigger"] = False

            if EX("t4") or EX("t41"):
                r["t41_red"] = False
            if EX("t4") or EX("t42"):
                r["t42_red"] = False
            if EX("t4") or EX("t43"):
                r["t43_red"] = False
            if EX("t4") or EX("t44"):
                r["t44_red"] = False

            if EX("t2") or EX("dt"):
                r["t2_red"] = False

            if EX("tacm"):
                r["tacm_red"] = False

            if EX("gacm"):
                r["gacm_red"] = False

            if EX("gacm_p") or EX("dgacm_val") or EX("dgacm"):
                r["dgacm_red"] = False

            if EX("g1"):
                r["g1_red"] = False

            if EX("dg_pct"):
                r["dgp_red"] = False

            if EX("dg"):
                r["dg_flow_red"] = False

            if EX("g_adaos"):
                r["gadaos_red"] = False

            if EX("time"):
                r["dataora_red"] = False
    else:
        for r in data:
            r["excluded_all"] = False

    filtered = data
    if any_filter_enabled:
        # === ЕСЛИ ЕСТЬ ХОТЬ ОДНО ВКЛЮЧЁННОЕ УСЛОВИЕ ===

        if season in ("Iarna", "Vara"):
            # Iarnă / Vară — работаем как раньше:
            # фильтруем строки и оставляем только те, что попали под условия
            filtered = []
            keep_for_search_ptc: set[str] = set()
            covered_by_params_ptc: set[str] = set()

            for r in data:
                ptc = str(r.get("ptc", ""))
                exset = active.get(ptc, set())
                red_before = before_map.get(ptc, set())

                if "__all__" in exset:
                    keep_for_search_ptc.add(ptc)
                    continue

                if red_before:
                    all_covered = all(_is_excluded(k, exset) for k in red_before)
                    if all_covered:
                        keep_for_search_ptc.add(ptc)
                        covered_by_params_ptc.add(ptc)

            for r in data:
                ptc = str(r.get("ptc", ""))
                r["excluded_by_params"] = ptc in covered_by_params_ptc

            for r in data:
                if (
                        (t1_en and r["t1_trigger"])
                        or (t4_en and (r["t41_red"] or r["t42_red"] or r["t43_red"] or r["t44_red"]))
                        or (dt_en and r["t2_red"])
                        or (tacm_en and r["tacm_red"])
                        or (gacm_max_en and r["gacm_red"])
                        or (dgacm_en and r["dgacm_red"])
                        or (g1_min_en and r["g1_red"])
                        or (dgp_en and r["dgp_red"])
                        or (dg_flow_en and r["dg_flow_red"])
                        or (gadaos_en and r["gadaos_red"])
                        or (dataora_en and r["dataora_red"])
                        or (pompa_off_en and r.get("pompa_off"))
                        or (sursa_off_en and r.get("sursa_off"))
                ):
                    filtered.append(r)

            if keep_for_search_ptc:
                have = {str(x.get("ptc", "")) for x in filtered}
                extra = [
                    r for r in data
                    if str(r.get("ptc", "")) in keep_for_search_ptc
                       and str(r.get("ptc", "")) not in have
                ]
                filtered.extend(extra)

        else:
            # === Toate ===
            # ВАЖНО: НИЧЕГО НЕ ФИЛЬТРУЕМ ПО УСЛОВИЯМ, только подсветка красным.
            # Все 318 объектов идут в ответ, excluded_by_params НЕ используется.
            for r in data:
                r["excluded_by_params"] = False
            filtered = data

    else:
        # === НЕТ ВКЛЮЧЁННЫХ УСЛОВИЙ (все галочки сняты) ===
        for r in data:
            r["excluded_by_params"] = False

        if season in ("Iarna", "Vara"):
            # Iarnă / Vară — таблица должна быть пустой
            filtered = []
        else:
            # Toate — показываем все объекты, без красного (условий нет)
            filtered = data

    # Cleanup old cache entries
    now_ts = timezone.now().timestamp()
    for k in list(LR_MEMORY_CACHE.keys()):
        _, ts = LR_MEMORY_CACHE[k]
        if now_ts - ts > LR_CACHE_TTL:
            del LR_MEMORY_CACHE[k]

    return JsonResponse(filtered, safe=False)


def export_ptc_excel(request):
    """
    Экспорт текущей таблицы Monitoring PTC в Excel.

    ВАЖНО:
      - Используем тот же набор GET-параметров, что и api_ptc_data (season, флаги условий и т.п.).
      - Не дублируем логику фильтров: просто вызываем api_ptc_data(request),
        парсим его JSON и по этим данным строим Excel.
    """
    # --- ТЕМА (light/dark) ---
    theme = (request.GET.get("theme") or "light").strip().lower()
    if theme not in EXCEL_THEME_COLORS:
        theme = "light"
    colors = EXCEL_THEME_COLORS[theme]

    # Получаем те же данные, что и для таблицы
    json_response = api_ptc_data(request)
    if json_response.status_code != 200:
        # Если api_ptc_data вернул ошибку — отдаём её как есть
        return json_response

    try:
        data = json.loads(json_response.content.decode("utf-8"))
    except Exception:
        data = []

    # --- ДОПОЛНИТЕЛЬНО: фильтр по Raion, как на фронте ---
    # radio-кнопки: Toate / 1 / 2 / 3 / 4 / 5
    raion = (request.GET.get("raion") or "").strip().lower()
    if raion and raion != "toate":

        def _match_raion(row: dict) -> bool:
            ptc_str = str(row.get("ptc") or "").strip()
            if not ptc_str:
                return False
            # номер района — первая цифра PTC
            first = ptc_str[0]
            return first == raion

        data = [row for row in data if _match_raion(row)]

    # --- Описание колонок (ключ в dict → заголовок в Excel) ---
    columns = [
        ("ptc", "PTC"),
        ("address", "Adresa"),
        ("q1", "Q"),
        ("g1", "G1"),
        ("g2", "G2"),
        ("dg", "ΔG"),
        ("dg_pct", "Δ%"),
        ("t1", "T1"),
        ("t2", "T2"),
        ("dt", "ΔT"),
        ("t31", "T31"),
        ("t32", "T32"),
        ("t41", "T41"),
        ("t42", "T42"),
        ("t43", "T43"),
        ("t44", "T44"),
        ("tacm", "Tacm"),
        ("gacm", "Gacm"),
        ("gacm_p", "Gacm-P"),
        ("dgacm_val", "ΔGacm"),
        ("g_adaos", "Gadaos"),
        ("sursa", "220V"),
        ("pompa", "Pompa"),
        ("time", "Data/Ora"),
    ]

    # Ограничение колонок по чекбоксам (если ты уже передаёшь ?cols=...)
    cols_param = (request.GET.get("cols") or "").strip()
    if cols_param:
        requested = {c.strip() for c in cols_param.split(",") if c.strip()}
        columns = [c for c in columns if c[0] in requested]

    # --- Создание Excel-файла ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Date"

    # Заголовок (первая строка)
    header_fill = PatternFill("solid", fgColor=colors["header_bg"])
    for col_idx, (_, title) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = Font(bold=True, color=colors["header_font"])
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    bullet = "●"  # символ для 220V

    # Данные
    for row in data:
        excel_row = []
        for key, _ in columns:

            # --- колонка 220V ---
            if key == "sursa":
                # кружок рисуем, если вообще есть объект 220V (id_sursa не пустой/не 0)
                id_sursa = row.get("id_sursa")
                has_220_object = id_sursa not in (None, "", 0, "0")
                excel_row.append(bullet if has_220_object else "")

            # --- колонка Pompa ---
            elif key == "pompa":
                vals = row.get("pompa")
                if isinstance(vals, list) and vals:
                    # рисуем по одному квадратику на каждую помпу
                    square = "■"
                    excel_row.append("".join(square for _ in vals))
                else:
                    excel_row.append("")

            # --- все остальные колонки без изменений ---
            else:
                val = row.get(key)
                excel_row.append(val)

        ws.append(excel_row)

        # --- стилизация только что добавленной строки ---
        row_idx = ws.max_row
        is_even = (row_idx % 2 == 0)
        base_fill = PatternFill(
            "solid",
            fgColor=colors["row_even_bg" if is_even else "row_odd_bg"],
        )

        for col_idx, (key, _) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = base_fill

            # базовый цвет текста
            cell.font = Font(color=colors["text"])

            # ---- подсветка по флагам, как в таблице ----
            # T1
            if key == "t1" and row.get("t1_trigger"):
                cell.font = Font(color=colors["danger"])

            # T2 (ΔT min)
            elif key == "t2" and row.get("t2_red"):
                cell.font = Font(color=colors["danger"])

            # T41..T44 (T4 min)
            elif key in ("t41", "t42", "t43", "t44") and row.get(f"{key}_red"):
                cell.font = Font(color=colors["danger"])

            # Tacm
            elif key == "tacm" and row.get("tacm_red"):
                cell.font = Font(color=colors["danger"])

            # Gacm
            elif key == "gacm" and row.get("gacm_red"):
                cell.font = Font(color=colors["danger"])

            # ΔGacm
            elif key == "dgacm_val" and row.get("dgacm_red"):
                cell.font = Font(color=colors["danger"])

            # G1 min
            elif key == "g1" and row.get("g1_red"):
                cell.font = Font(color=colors["danger"])

            # Δ% (dgp_red)
            elif key == "dg_pct" and row.get("dgp_red"):
                cell.font = Font(color=colors["danger"])

            # ΔG (dg_flow_red)
            elif key == "dg" and row.get("dg_flow_red"):
                cell.font = Font(color=colors["danger"])

            # Gadaos
            elif key == "g_adaos" and row.get("gadaos_red"):
                cell.font = Font(color=colors["danger"])

            # Data/Ora
            elif key == "time" and row.get("dataora_red"):
                cell.font = Font(color=colors["danger"])

            # 220V (sursa): зелёный/красный кружок
            elif key == "sursa":
                val = row.get("sursa")
                if val:
                    # есть 220V
                    cell.font = Font(color=colors["good"])
                elif row.get("id_sursa"):
                    # объект есть, но 220V нет → красный
                    cell.font = Font(color=colors["voltage_off"])

            # Pompa: если список не пустой — сделаем зелёным
            elif key == "pompa":
                vals = row.get("pompa") or []
                if isinstance(vals, list) and vals:
                    lcs = row.get("lcs") or 0
                    LCS_NORM = 30

                    # определяем цвет по той же логике, что и в JS
                    if lcs < LCS_NORM:
                        pump_color = colors["pump_yellow"]
                    else:
                        has_red = any((v or 0) > 200 for v in vals)
                        has_green = any((v or 0) > 0 for v in vals)

                        if has_red:
                            pump_color = colors["danger"]
                        elif has_green:
                            pump_color = colors["good"]
                        else:
                            pump_color = colors["pump_gray"]

                    cell.font = Font(color=pump_color)
                # если списка помп нет — оставляем базовый цвет текста

    # Автоширина колонок
    for col_idx, _ in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            text = str(cell.value) if cell.value is not None else ""
            if len(text) > max_len:
                max_len = len(text)
        ws.column_dimensions[col_letter].width = max(8, min(max_len + 2, 40))

    # --- HTTP-ответ с Excel-файлом ---
    filename = f"monitoring_ptc_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ───────── Дополнительные Django Views ─────────
def exclude_view(request, ptc: str):
    """Страница исключений для PTC."""
    exclusions = _load_json(EXCLUSIONS_PATH)
    existing = exclusions.get(ptc, [])

    possible_params = [
        "PTC ",
        "Q",
        "G1",
        "G2",
        "ΔG",
        "Δ%",
        "T1",
        "T2",
        "ΔT",
        "T31",
        "T32",
        "T41",
        "T42",
        "T43",
        "T44",
        "Tacm",
        "Gacm",
        "Gacm-P",
        "ΔGacm",
        "Gadaos",
        "220V",
        "Pompa",
        "Data/Ora",
    ]

    can_edit = can_edit_from_request(request)

    if request.method == "POST":
        if not can_edit:
            return HttpResponseForbidden("Редактирование запрещено для вашего IP.")

        action = request.POST.get("action")
        if action == "add":
            chosen_label = (request.POST.get("param") or "").strip()
            keys = _norm_excl_param(chosen_label)

            # Read start and end of the exclusion period
            start = request.POST.get("start")
            end = request.POST.get("end") or request.POST.get("until")

            tura = (request.POST.get("tura") or "4")
            reason = request.POST.get("reason") or ""

            # сохраняем момент добавления в локальном часовом поясе Europe/Chisinau,
            # fallback на timezone.now() (UTC) при ошибке
            try:
                tz_local = ZoneInfo(settings.LOCAL_TIME_ZONE)
            except Exception:
                tz_local = None
            if tz_local is not None:
                ts = datetime.now(tz_local).isoformat(timespec="seconds")
            else:
                ts = timezone.now().isoformat(timespec="seconds")

            new_items = existing[:]
            for k in keys:
                item = {
                    "param": k,
                    # Save start and end of exclusion period
                    "start": start,
                    "end": end,
                    "tura": tura,
                    "reason": reason,
                    "ts": ts,
                }
                # Optional: persist the date portion of the end as 'until'
                if end:
                    try:
                        item["until"] = end.split("T")[0]
                    except Exception:
                        item["until"] = end
                new_items.append(item)

            exclusions[ptc] = new_items
            _save_json(EXCLUSIONS_PATH, exclusions)
            return HttpResponseRedirect(request.path)

        elif action == "delete":
            idx = int(request.POST.get("index", -1))
            if 0 <= idx < len(existing):
                del existing[idx]
                exclusions[ptc] = existing
                _save_json(EXCLUSIONS_PATH, exclusions)
            return HttpResponseRedirect(request.path)

    tomorrow = timezone.localdate() + timedelta(days=1)

    return render(
        request,
        "monitoring/exclude_page.html",
        {
            "ptc": ptc,
            "exclusions": existing,
            "params": possible_params,
            "tomorrow_iso": tomorrow.isoformat(),
            "can_edit": can_edit,
        },
    )


def comment_view(request, ptc: str):
    """Страница комментариев для PTC (add + delete)."""
    comments = _load_json(COMMENTS_PATH)
    existing = comments.get(ptc, [])

    can_edit = can_edit_from_request(request)

    if request.method == "POST":
        if not can_edit:
            return HttpResponseForbidden("Редактирование запрещено для вашего IP.")

        action = (request.POST.get("action") or "add").lower()

        if action == "add":
            text = (request.POST.get("text") or "").strip()
            if text:
                ts = timezone.now().isoformat(timespec="seconds")
                rec = {"text": text, "ts": ts}
                comments[ptc] = existing + [rec]
                _save_json(COMMENTS_PATH, comments)
            return HttpResponseRedirect(request.path)

        elif action == "delete":
            try:
                idx = int(request.POST.get("index"))
            except (TypeError, ValueError):
                idx = -1
            if 0 <= idx < len(existing):
                del existing[idx]
                comments[ptc] = existing
                _save_json(COMMENTS_PATH, comments)
            return HttpResponseRedirect(request.path)

    return render(
        request,
        "monitoring/comment_page.html",
        {
            "ptc": ptc,
            "comments": existing,
            "can_edit": can_edit,
        },
    )
